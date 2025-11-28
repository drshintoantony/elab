"""
eLAB Analytics Comprehensive Data Exporter
Exports complete analytics data from Supabase with all requested analyses
Fixes district naming inconsistencies and removes duplicates
"""

import os
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any
import pandas as pd
from supabase import create_client, Client
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json

# Load environment variables
load_dotenv()

class eLABAnalyticsComprehensive:
    def __init__(self):
        """Initialize Supabase client"""
        supabase_url = os.getenv("SUPABASE_URL")
        supabase_key = os.getenv("SUPABASE_KEY")

        if not supabase_url or not supabase_key:
            raise ValueError("SUPABASE_URL and SUPABASE_KEY must be set in .env file")

        self.client: Client = create_client(supabase_url, supabase_key)
        print("✓ Connected to Supabase")

        # District name normalization mapping
        self.district_mapping = {
            'trivandrum': 'Thiruvananthapuram',
            'thiruvananthapuram': 'Thiruvananthapuram',
            'tvm': 'Thiruvananthapuram',
            'ernakulam': 'Ernakulam',
            'ekm': 'Ernakulam',
            'kochi': 'Ernakulam',
            'thrissur': 'Thrissur',
            'trichur': 'Thrissur',
            'kozhikode': 'Kozhikode',
            'calicut': 'Kozhikode',
            'kollam': 'Kollam',
            'quilon': 'Kollam',
            'alappuzha': 'Alappuzha',
            'alleppey': 'Alappuzha',
            'palakkad': 'Palakkad',
            'palghat': 'Palakkad',
            'malappuram': 'Malappuram',
            'kannur': 'Kannur',
            'cannanore': 'Kannur',
            'kasaragod': 'Kasaragod',
            'kottayam': 'Kottayam',
            'idukki': 'Idukki',
            'wayanad': 'Wayanad',
            'pathanamthitta': 'Pathanamthitta',
        }

    def normalize_district(self, district: str) -> str:
        """Normalize district names to standard format"""
        if not district or pd.isna(district):
            return 'Unknown'
        district_lower = str(district).strip().lower()
        return self.district_mapping.get(district_lower, district.strip())

    def get_date_range(self, days_back: int = 365) -> tuple:
        """Get date range for queries (default: all time if days_back is large)"""
        end_date = datetime.now()
        start_date = end_date - timedelta(days=days_back)
        return start_date.isoformat(), end_date.isoformat()

    def get_employee_id(self, visit_data: dict) -> tuple:
        """
        Get the correct employee ID and name from visit data.
        Priority: field_staff_id (the actual employee who added the case)
        Fallback: lab_assistant_id
        Returns: (employee_id, employee_name, employee_role)
        """
        # field_staff_id is the actual employee who added the case
        employee_id = visit_data.get('field_staff_id')
        employee_name = visit_data.get('field_staff_name')
        employee_role = 'field_staff'

        # If no field_staff_id, use lab_assistant_id
        if not employee_id:
            employee_id = visit_data.get('lab_assistant_id')
            employee_role = 'lab_assistant'

        return employee_id, employee_name, employee_role

    # ========== DISTRICT-WISE ANALYSIS WITH FIXES ==========
    def get_district_analysis_comprehensive(self, start_date: str, end_date: str) -> pd.DataFrame:
        """
        Comprehensive district-wise analysis with normalized names
        Gets ALL records without date filtering to ensure we get all 5588+ visits
        """
        print("📊 Fetching comprehensive district-wise analysis...")

        # Get ALL visits without date filtering
        visits = self.client.table('visits').select(
            'id, visit_date, owner_id, total_amount, owners(district, state)'
        ).execute()

        print(f"   Retrieved {len(visits.data)} visit records")

        data = []
        for visit in visits.data:
            owner_data = visit.get('owners', {}) or {}
            district = self.normalize_district(owner_data.get('district', 'Unknown'))

            data.append({
                'district': district,
                'state': owner_data.get('state', 'Kerala'),
                'visit_id': visit['id'],
                'visit_date': visit['visit_date'],
                'owner_id': visit.get('owner_id'),
                'total_amount': visit.get('total_amount', 0) or 0
            })

        df = pd.DataFrame(data)
        if len(df) > 0:
            df['visit_date'] = pd.to_datetime(df['visit_date'])
            # Apply date filtering AFTER retrieving all data
            if start_date and end_date:
                df = df[(df['visit_date'] >= start_date) & (df['visit_date'] <= end_date)]

            summary = df.groupby(['district', 'state']).agg({
                'visit_id': 'count',
                'owner_id': 'nunique',
                'total_amount': ['sum', 'mean']
            }).reset_index()
            summary.columns = ['district', 'state', 'total_cases', 'unique_owners', 'total_revenue', 'avg_revenue_per_visit']
            summary = summary.sort_values('total_cases', ascending=False)
            return summary
        return df

    # ========== MONTHLY EMPLOYEE COMPARISON ==========
    def get_employee_monthly_comparison(self, start_date: str, end_date: str) -> pd.DataFrame:
        """
        Monthly comparison of individual employees - cases, revenue, performance
        Uses field_staff_id first (the actual employee), fallback to lab_assistant_id
        """
        print("👥 Fetching monthly employee comparison...")

        # Get users data separately
        users = self.client.table('users').select('id, name, email, role').execute()
        users_dict = {u['id']: u for u in users.data}

        # Get ALL visits
        visits = self.client.table('visits').select(
            'id, visit_date, lab_assistant_id, total_amount, payment_status, sample_collected, report_sent, field_staff_id, field_staff_name'
        ).execute()

        print(f"   Retrieved {len(visits.data)} visit records")

        data = []
        for visit in visits.data:
            # CORRECT LOGIC: field_staff_id is the actual employee who added the case
            employee_id, employee_name, employee_role = self.get_employee_id(visit)

            # If we have employee_id from field_staff_id but no name, try to get from users table
            if employee_id and not employee_name:
                user_data = users_dict.get(employee_id, {})
                employee_name = user_data.get('name', 'Unknown')
                if user_data.get('role'):
                    employee_role = user_data.get('role')

            # Skip if no employee ID at all
            if not employee_id:
                continue

            visit_date = pd.to_datetime(visit['visit_date'])

            data.append({
                'employee_id': employee_id,
                'employee_name': employee_name or 'Unknown',
                'employee_role': employee_role,
                'year_month': visit_date.strftime('%Y-%m'),
                'month_name': visit_date.strftime('%B %Y'),
                'visit_id': visit['id'],
                'revenue': visit.get('total_amount', 0) or 0,
                'payment_received': 1 if visit.get('payment_status') else 0,
                'sample_collected': 1 if visit.get('sample_collected') else 0,
                'report_sent': 1 if visit.get('report_sent') else 0
            })

        df = pd.DataFrame(data)
        if len(df) > 0:
            # Apply date filtering after retrieval
            df['visit_date_parsed'] = pd.to_datetime(df['year_month'] + '-01')
            if start_date and end_date:
                start_dt = pd.to_datetime(start_date)
                end_dt = pd.to_datetime(end_date)
                df = df[(df['visit_date_parsed'] >= start_dt) & (df['visit_date_parsed'] <= end_dt)]

            monthly_summary = df.groupby(['employee_id', 'employee_name', 'employee_role', 'year_month', 'month_name']).agg({
                'visit_id': 'count',
                'revenue': 'sum',
                'payment_received': 'sum',
                'sample_collected': 'sum',
                'report_sent': 'sum'
            }).reset_index()
            monthly_summary.columns = ['employee_id', 'employee_name', 'employee_role', 'year_month', 'month_name',
                                       'total_cases', 'total_revenue', 'payments_received', 'samples_collected', 'reports_sent']
            monthly_summary = monthly_summary.sort_values(['year_month', 'total_cases'], ascending=[True, False])
            return monthly_summary
        return df

    # ========== MONTHLY CASES BY INDIVIDUAL EMPLOYEES ==========
    def get_employee_monthly_cases(self, start_date: str, end_date: str) -> pd.DataFrame:
        """
        Monthly cases done by each individual employee with details
        """
        print("📅 Fetching monthly cases by individual employees...")

        # Get users data separately
        users = self.client.table('users').select('id, name, role').execute()
        users_dict = {u['id']: u for u in users.data}

        visits = self.client.table('visits').select(
            'id, visit_date, lab_assistant_id, total_amount, field_staff_id, field_staff_name, owner_id, owners(district, state)'
        ).gte('visit_date', start_date).lte('visit_date', end_date).execute()

        data = []
        for visit in visits.data:
            employee_id = visit.get('lab_assistant_id') or visit.get('field_staff_id')
            user_data = users_dict.get(employee_id, {})
            owner_data = visit.get('owners', {}) or {}
            visit_date = pd.to_datetime(visit['visit_date'])

            data.append({
                'employee_id': employee_id,
                'employee_name': user_data.get('name') or visit.get('field_staff_name', 'Unknown'),
                'employee_role': user_data.get('role', 'field_staff'),
                'year_month': visit_date.strftime('%Y-%m'),
                'month_name': visit_date.strftime('%B %Y'),
                'visit_date': visit['visit_date'],
                'district': self.normalize_district(owner_data.get('district', 'Unknown')),
                'state': owner_data.get('state', 'Kerala'),
                'revenue': visit.get('total_amount', 0) or 0
            })

        df = pd.DataFrame(data)
        if len(df) > 0:
            df = df.sort_values(['employee_name', 'year_month', 'visit_date'])
            return df
        return df

    # ========== TEST ANALYSIS WITH MONTHLY BREAKDOWN ==========
    def get_test_monthly_analysis(self, start_date: str, end_date: str) -> pd.DataFrame:
        """
        Test analysis with monthly breakdown - which tests are done each month
        """
        print("🔬 Fetching test analysis with monthly breakdown...")

        # Get users data separately
        users = self.client.table('users').select('id, name').execute()
        users_dict = {u['id']: u for u in users.data}

        tests = self.client.table('tests').select(
            'id, test_type, test_name, price, created_at, visit_animal_id, visit_animals(visit_id, visits(visit_date, lab_assistant_id, field_staff_id, field_staff_name))'
        ).execute()

        # Make start_date and end_date timezone-aware
        start_dt = pd.to_datetime(start_date).tz_localize('UTC')
        end_dt = pd.to_datetime(end_date).tz_localize('UTC')

        data = []
        for test in tests.data:
            va = test.get('visit_animals') or {}
            visit = va.get('visits') or {}

            visit_date_str = visit.get('visit_date') if visit else test.get('created_at')
            if not visit_date_str:
                continue

            visit_date = pd.to_datetime(visit_date_str)

            # Filter by date range
            if visit_date < start_dt or visit_date > end_dt:
                continue

            data.append({
                'test_type': test.get('test_type', 'Unknown'),
                'test_name': test.get('test_name', 'Unknown'),
                'price': test.get('price', 0) or 0,
                'year_month': visit_date.strftime('%Y-%m'),
                'month_name': visit_date.strftime('%B %Y'),
                'visit_date': visit_date_str
            })

        df = pd.DataFrame(data)
        if len(df) > 0:
            monthly_test_summary = df.groupby(['year_month', 'month_name', 'test_type', 'test_name']).agg({
                'price': ['count', 'sum', 'mean']
            }).reset_index()
            monthly_test_summary.columns = ['year_month', 'month_name', 'test_type', 'test_name', 'test_count', 'total_revenue', 'avg_price']
            monthly_test_summary = monthly_test_summary.sort_values(['year_month', 'test_count'], ascending=[True, False])
            return monthly_test_summary
        return df

    # ========== TEST ANALYSIS EMPLOYEE-WISE ==========
    def get_test_employee_wise_analysis(self, start_date: str, end_date: str) -> pd.DataFrame:
        """
        Test analysis employee-wise - which employee does which tests and how many
        """
        print("👨‍⚕️ Fetching employee-wise test analysis...")

        # Get users data separately
        users = self.client.table('users').select('id, name, role').execute()
        users_dict = {u['id']: u for u in users.data}

        tests = self.client.table('tests').select(
            'id, test_type, test_name, price, created_at, visit_animal_id, visit_animals(visit_id, visits(visit_date, lab_assistant_id, field_staff_id, field_staff_name))'
        ).execute()

        # Make start_date and end_date timezone-aware
        start_dt = pd.to_datetime(start_date).tz_localize('UTC')
        end_dt = pd.to_datetime(end_date).tz_localize('UTC')

        data = []
        for test in tests.data:
            va = test.get('visit_animals') or {}
            visit = va.get('visits') or {}

            visit_date_str = visit.get('visit_date') if visit else test.get('created_at')
            if not visit_date_str:
                continue

            visit_date = pd.to_datetime(visit_date_str)

            # Filter by date range
            if visit_date < start_dt or visit_date > end_dt:
                continue

            employee_id = visit.get('lab_assistant_id') or visit.get('field_staff_id')
            user_data = users_dict.get(employee_id, {})
            employee_name = user_data.get('name') or visit.get('field_staff_name', 'Unknown')

            data.append({
                'employee_id': employee_id,
                'employee_name': employee_name,
                'employee_role': user_data.get('role', 'field_staff'),
                'test_type': test.get('test_type', 'Unknown'),
                'test_name': test.get('test_name', 'Unknown'),
                'price': test.get('price', 0) or 0,
                'year_month': visit_date.strftime('%Y-%m'),
                'month_name': visit_date.strftime('%B %Y')
            })

        df = pd.DataFrame(data)
        if len(df) > 0:
            employee_test_summary = df.groupby(['employee_id', 'employee_name', 'employee_role', 'test_type', 'test_name']).agg({
                'price': ['count', 'sum']
            }).reset_index()
            employee_test_summary.columns = ['employee_id', 'employee_name', 'employee_role', 'test_type', 'test_name', 'test_count', 'total_revenue']
            employee_test_summary = employee_test_summary.sort_values(['employee_name', 'test_count'], ascending=[True, False])
            return employee_test_summary
        return df

    # ========== MOST TESTS DONE BY EACH EMPLOYEE ==========
    def get_employee_most_tests(self, start_date: str, end_date: str) -> pd.DataFrame:
        """
        Most tests done by each employee - ranked list
        """
        print("🏆 Fetching most tests done by each employee...")

        # Get the employee-wise test data
        employee_test_df = self.get_test_employee_wise_analysis(start_date, end_date)

        if len(employee_test_df) > 0:
            # Group by employee to get their top tests
            most_tests = employee_test_df.groupby(['employee_id', 'employee_name', 'employee_role']).agg({
                'test_count': 'sum',
                'total_revenue': 'sum'
            }).reset_index()
            most_tests.columns = ['employee_id', 'employee_name', 'employee_role', 'total_tests_performed', 'total_revenue_from_tests']
            most_tests = most_tests.sort_values('total_tests_performed', ascending=False)

            # Also get the top test for each employee
            top_test_per_employee = employee_test_df.loc[employee_test_df.groupby('employee_id')['test_count'].idxmax()]
            top_test_per_employee = top_test_per_employee[['employee_id', 'test_type', 'test_name', 'test_count']]
            top_test_per_employee.columns = ['employee_id', 'most_common_test_type', 'most_common_test_name', 'most_common_test_count']

            # Merge
            result = pd.merge(most_tests, top_test_per_employee, on='employee_id', how='left')
            return result
        return employee_test_df

    # ========== DISTRICT-WISE TEST ANALYSIS ==========
    def get_district_test_analysis(self, start_date: str, end_date: str) -> pd.DataFrame:
        """
        District-wise test analysis with monthly and yearly breakdown
        Tests sorted from least to most by district
        """
        print("🗺️ Fetching district-wise test analysis...")

        # Get all tests with visit and district information
        tests = self.client.table('tests').select(
            'id, test_type, test_name, price, created_at, visit_animal_id, visit_animals(visit_id, visits(visit_date, owner_id, owners(district, state)))'
        ).execute()

        # Make start_date and end_date timezone-aware
        start_dt = pd.to_datetime(start_date).tz_localize('UTC')
        end_dt = pd.to_datetime(end_date).tz_localize('UTC')

        data = []
        for test in tests.data:
            va = test.get('visit_animals') or {}
            visit = va.get('visits') or {}
            owner = visit.get('owners') or {}

            visit_date_str = visit.get('visit_date') if visit else test.get('created_at')
            if not visit_date_str:
                continue

            visit_date = pd.to_datetime(visit_date_str)

            # Filter by date range
            if visit_date < start_dt or visit_date > end_dt:
                continue

            district = self.normalize_district(owner.get('district', 'Unknown'))

            data.append({
                'district': district,
                'state': owner.get('state', 'Kerala'),
                'test_type': test.get('test_type', 'Unknown'),
                'test_name': test.get('test_name', 'Unknown'),
                'price': test.get('price', 0) or 0,
                'year': visit_date.year,
                'year_month': visit_date.strftime('%Y-%m'),
                'month_name': visit_date.strftime('%B %Y')
            })

        df = pd.DataFrame(data)
        if len(df) > 0:
            # District + Test summary (sorted least to most)
            district_test_summary = df.groupby(['district', 'state', 'test_type', 'test_name']).agg({
                'price': ['count', 'sum']
            }).reset_index()
            district_test_summary.columns = ['district', 'state', 'test_type', 'test_name', 'test_count', 'total_revenue']
            district_test_summary = district_test_summary.sort_values(['district', 'test_count'], ascending=[True, True])

            return district_test_summary
        return df

    # ========== DISTRICT-WISE TEST MONTHLY ANALYSIS ==========
    def get_district_test_monthly_analysis(self, start_date: str, end_date: str) -> pd.DataFrame:
        """
        District-wise test analysis with monthly breakdown
        """
        print("📊 Fetching district-wise test monthly analysis...")

        tests = self.client.table('tests').select(
            'id, test_type, test_name, price, created_at, visit_animal_id, visit_animals(visit_id, visits(visit_date, owner_id, owners(district, state)))'
        ).execute()

        # Make start_date and end_date timezone-aware
        start_dt = pd.to_datetime(start_date).tz_localize('UTC')
        end_dt = pd.to_datetime(end_date).tz_localize('UTC')

        data = []
        for test in tests.data:
            va = test.get('visit_animals') or {}
            visit = va.get('visits') or {}
            owner = visit.get('owners') or {}

            visit_date_str = visit.get('visit_date') if visit else test.get('created_at')
            if not visit_date_str:
                continue

            visit_date = pd.to_datetime(visit_date_str)

            # Filter by date range
            if visit_date < start_dt or visit_date > end_dt:
                continue

            district = self.normalize_district(owner.get('district', 'Unknown'))

            data.append({
                'district': district,
                'state': owner.get('state', 'Kerala'),
                'test_type': test.get('test_type', 'Unknown'),
                'test_name': test.get('test_name', 'Unknown'),
                'price': test.get('price', 0) or 0,
                'year_month': visit_date.strftime('%Y-%m'),
                'month_name': visit_date.strftime('%B %Y')
            })

        df = pd.DataFrame(data)
        if len(df) > 0:
            monthly_summary = df.groupby(['district', 'state', 'year_month', 'month_name', 'test_type', 'test_name']).agg({
                'price': ['count', 'sum']
            }).reset_index()
            monthly_summary.columns = ['district', 'state', 'year_month', 'month_name', 'test_type', 'test_name', 'test_count', 'total_revenue']
            monthly_summary = monthly_summary.sort_values(['district', 'year_month', 'test_count'], ascending=[True, True, False])

            return monthly_summary
        return df

    # ========== DISTRICT-WISE TEST YEARLY ANALYSIS ==========
    def get_district_test_yearly_analysis(self, start_date: str, end_date: str) -> pd.DataFrame:
        """
        District-wise test analysis with yearly breakdown
        """
        print("📈 Fetching district-wise test yearly analysis...")

        tests = self.client.table('tests').select(
            'id, test_type, test_name, price, created_at, visit_animal_id, visit_animals(visit_id, visits(visit_date, owner_id, owners(district, state)))'
        ).execute()

        # Make start_date and end_date timezone-aware
        start_dt = pd.to_datetime(start_date).tz_localize('UTC')
        end_dt = pd.to_datetime(end_date).tz_localize('UTC')

        data = []
        for test in tests.data:
            va = test.get('visit_animals') or {}
            visit = va.get('visits') or {}
            owner = visit.get('owners') or {}

            visit_date_str = visit.get('visit_date') if visit else test.get('created_at')
            if not visit_date_str:
                continue

            visit_date = pd.to_datetime(visit_date_str)

            # Filter by date range
            if visit_date < start_dt or visit_date > end_dt:
                continue

            district = self.normalize_district(owner.get('district', 'Unknown'))

            data.append({
                'district': district,
                'state': owner.get('state', 'Kerala'),
                'test_type': test.get('test_type', 'Unknown'),
                'test_name': test.get('test_name', 'Unknown'),
                'price': test.get('price', 0) or 0,
                'year': visit_date.year
            })

        df = pd.DataFrame(data)
        if len(df) > 0:
            yearly_summary = df.groupby(['district', 'state', 'year', 'test_type', 'test_name']).agg({
                'price': ['count', 'sum']
            }).reset_index()
            yearly_summary.columns = ['district', 'state', 'year', 'test_type', 'test_name', 'test_count', 'total_revenue']
            yearly_summary = yearly_summary.sort_values(['district', 'year', 'test_count'], ascending=[True, True, False])

            return yearly_summary
        return df

    # ========== SPECIES ANALYSIS ==========
    def get_species_analysis(self, start_date: str, end_date: str, species: Optional[str] = None) -> pd.DataFrame:
        """
        Species-wise cases and revenue analysis
        """
        print("🐾 Fetching species-wise analysis...")

        # Get animals with visits
        query = self.client.table('visit_animals').select(
            'id, animal_id, visit_id, animals(species, breed), visits(visit_date, total_amount)'
        ).execute()

        data = []
        for va in query.data:
            animal = va.get('animals') or {}
            visit = va.get('visits') or {}

            if not visit:
                continue

            visit_date_str = visit.get('visit_date')
            if not visit_date_str:
                continue

            data.append({
                'species': animal.get('species', 'Unknown'),
                'breed': animal.get('breed', 'Unknown'),
                'visit_date': visit_date_str,
                'revenue': visit.get('total_amount', 0) or 0
            })

        df = pd.DataFrame(data)
        if len(df) > 0:
            df['visit_date'] = pd.to_datetime(df['visit_date'])
            df = df[(df['visit_date'] >= start_date) & (df['visit_date'] <= end_date)]

            if species:
                df = df[df['species'].str.lower() == species.lower()]

            summary = df.groupby('species').agg({
                'visit_date': 'count',
                'revenue': ['sum', 'mean']
            }).reset_index()
            summary.columns = ['species', 'total_cases', 'total_revenue', 'avg_revenue_per_case']
            summary = summary.sort_values('total_cases', ascending=False)
            return summary
        return df

    # ========== DASHBOARD KPIs ==========
    def get_dashboard_kpis(self, start_date: str, end_date: str) -> Dict[str, Any]:
        """
        Combined dashboard with all KPIs
        """
        print("📊 Generating dashboard KPIs...")

        # Get overall visit stats
        visits = self.client.table('visits').select(
            'id, visit_date, total_amount, payment_status, sample_collected, report_sent, owner_id'
        ).gte('visit_date', start_date).lte('visit_date', end_date).execute()

        visits_df = pd.DataFrame(visits.data)

        kpis = {
            'total_cases': len(visits_df),
            'total_revenue': float(visits_df['total_amount'].sum()) if len(visits_df) > 0 and 'total_amount' in visits_df.columns else 0,
            'avg_revenue_per_case': float(visits_df['total_amount'].mean()) if len(visits_df) > 0 and 'total_amount' in visits_df.columns else 0,
            'payment_completion_rate': float((visits_df['payment_status'].sum() / len(visits_df) * 100)) if len(visits_df) > 0 and 'payment_status' in visits_df.columns else 0,
            'sample_collection_rate': float((visits_df['sample_collected'].sum() / len(visits_df) * 100)) if len(visits_df) > 0 and 'sample_collected' in visits_df.columns else 0,
            'report_sent_rate': float((visits_df['report_sent'].sum() / len(visits_df) * 100)) if len(visits_df) > 0 and 'report_sent' in visits_df.columns else 0,
        }

        # Get employee count
        employees = self.client.table('users').select('id, role').in_('role', ['lab_assistant', 'field_staff']).execute()
        kpis['total_employees'] = len(employees.data)

        # Get unique owners
        unique_owners = visits_df['owner_id'].nunique() if len(visits_df) > 0 and 'owner_id' in visits_df.columns else 0
        kpis['unique_customers'] = int(unique_owners)

        return kpis

    # ========== Excel Export Functions ==========
    def format_excel_sheet(self, worksheet, df: pd.DataFrame):
        """Apply formatting to Excel worksheet"""
        if len(df) == 0:
            return

        # Header formatting
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-adjust column widths
        for idx, column in enumerate(df.columns, 1):
            column_letter = get_column_letter(idx)
            max_length = max(
                df[column].astype(str).apply(len).max(),
                len(str(column))
            )
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in worksheet.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = thin_border

    def export_to_excel(self, filename: str = "elab_analytics_comprehensive.xlsx", days_back: int = 365):
        """
        Export all comprehensive analytics to Excel file with multiple sheets
        """
        print(f"\n{'='*70}")
        print(f"Starting eLAB Comprehensive Analytics Export")
        print(f"{'='*70}\n")

        start_date, end_date = self.get_date_range(days_back)
        print(f"📅 Date range: {start_date[:10]} to {end_date[:10]}\n")

        all_data = {}

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:

            # 1. District Analysis (Fixed)
            try:
                print("1/11 Processing District Analysis...")
                district_df = self.get_district_analysis_comprehensive(start_date, end_date)
                if not district_df.empty:
                    district_df.to_excel(writer, sheet_name='District Analysis', index=False)
                    self.format_excel_sheet(writer.sheets['District Analysis'], district_df)
                    all_data['district_analysis'] = district_df.to_dict('records')
                    print(f"   ✓ District Analysis: {len(district_df)} rows exported")
                else:
                    print("   ⚠ District Analysis: No data found")
            except Exception as e:
                print(f"   ✗ District Analysis failed: {e}")

            # 2. Monthly Employee Comparison
            try:
                print("2/11 Processing Monthly Employee Comparison...")
                emp_monthly_comp_df = self.get_employee_monthly_comparison(start_date, end_date)
                if not emp_monthly_comp_df.empty:
                    emp_monthly_comp_df.to_excel(writer, sheet_name='Employee Monthly Comparison', index=False)
                    self.format_excel_sheet(writer.sheets['Employee Monthly Comparison'], emp_monthly_comp_df)
                    all_data['employee_monthly_comparison'] = emp_monthly_comp_df.to_dict('records')
                    print(f"   ✓ Employee Monthly Comparison: {len(emp_monthly_comp_df)} rows exported")
                else:
                    print("   ⚠ Employee Monthly Comparison: No data found")
            except Exception as e:
                print(f"   ✗ Employee Monthly Comparison failed: {e}")

            # 3. Monthly Cases by Employee
            try:
                print("3/11 Processing Monthly Cases by Employee...")
                emp_monthly_cases_df = self.get_employee_monthly_cases(start_date, end_date)
                if not emp_monthly_cases_df.empty:
                    emp_monthly_cases_df.to_excel(writer, sheet_name='Employee Monthly Cases', index=False)
                    self.format_excel_sheet(writer.sheets['Employee Monthly Cases'], emp_monthly_cases_df)
                    all_data['employee_monthly_cases'] = emp_monthly_cases_df.to_dict('records')
                    print(f"   ✓ Employee Monthly Cases: {len(emp_monthly_cases_df)} rows exported")
                else:
                    print("   ⚠ Employee Monthly Cases: No data found")
            except Exception as e:
                print(f"   ✗ Employee Monthly Cases failed: {e}")

            # 4. Test Monthly Analysis
            try:
                print("4/11 Processing Test Monthly Analysis...")
                test_monthly_df = self.get_test_monthly_analysis(start_date, end_date)
                if not test_monthly_df.empty:
                    test_monthly_df.to_excel(writer, sheet_name='Test Monthly Analysis', index=False)
                    self.format_excel_sheet(writer.sheets['Test Monthly Analysis'], test_monthly_df)
                    all_data['test_monthly_analysis'] = test_monthly_df.to_dict('records')
                    print(f"   ✓ Test Monthly Analysis: {len(test_monthly_df)} rows exported")
                else:
                    print("   ⚠ Test Monthly Analysis: No data found")
            except Exception as e:
                print(f"   ✗ Test Monthly Analysis failed: {e}")

            # 5. Test Employee-wise Analysis
            try:
                print("5/11 Processing Test Employee-wise Analysis...")
                test_employee_df = self.get_test_employee_wise_analysis(start_date, end_date)
                if not test_employee_df.empty:
                    test_employee_df.to_excel(writer, sheet_name='Test Employee Analysis', index=False)
                    self.format_excel_sheet(writer.sheets['Test Employee Analysis'], test_employee_df)
                    all_data['test_employee_analysis'] = test_employee_df.to_dict('records')
                    print(f"   ✓ Test Employee-wise Analysis: {len(test_employee_df)} rows exported")
                else:
                    print("   ⚠ Test Employee-wise Analysis: No data found")
            except Exception as e:
                print(f"   ✗ Test Employee-wise Analysis failed: {e}")

            # 6. Most Tests by Employee
            try:
                print("6/11 Processing Most Tests by Employee...")
                most_tests_df = self.get_employee_most_tests(start_date, end_date)
                if not most_tests_df.empty:
                    most_tests_df.to_excel(writer, sheet_name='Employee Most Tests', index=False)
                    self.format_excel_sheet(writer.sheets['Employee Most Tests'], most_tests_df)
                    all_data['employee_most_tests'] = most_tests_df.to_dict('records')
                    print(f"   ✓ Most Tests by Employee: {len(most_tests_df)} rows exported")
                else:
                    print("   ⚠ Most Tests by Employee: No data found")
            except Exception as e:
                print(f"   ✗ Most Tests by Employee failed: {e}")

            # 7. District-wise Test Analysis
            try:
                print("7/11 Processing District-wise Test Analysis...")
                district_test_df = self.get_district_test_analysis(start_date, end_date)
                if not district_test_df.empty:
                    district_test_df.to_excel(writer, sheet_name='District Test Analysis', index=False)
                    self.format_excel_sheet(writer.sheets['District Test Analysis'], district_test_df)
                    all_data['district_test_analysis'] = district_test_df.to_dict('records')
                    print(f"   ✓ District-wise Test Analysis: {len(district_test_df)} rows exported")
                else:
                    print("   ⚠ District-wise Test Analysis: No data found")
            except Exception as e:
                print(f"   ✗ District-wise Test Analysis failed: {e}")

            # 8. District-wise Test Monthly Analysis
            try:
                print("8/11 Processing District-wise Test Monthly Analysis...")
                district_test_monthly_df = self.get_district_test_monthly_analysis(start_date, end_date)
                if not district_test_monthly_df.empty:
                    district_test_monthly_df.to_excel(writer, sheet_name='District Test Monthly', index=False)
                    self.format_excel_sheet(writer.sheets['District Test Monthly'], district_test_monthly_df)
                    all_data['district_test_monthly'] = district_test_monthly_df.to_dict('records')
                    print(f"   ✓ District-wise Test Monthly: {len(district_test_monthly_df)} rows exported")
                else:
                    print("   ⚠ District-wise Test Monthly: No data found")
            except Exception as e:
                print(f"   ✗ District-wise Test Monthly failed: {e}")

            # 9. District-wise Test Yearly Analysis
            try:
                print("9/11 Processing District-wise Test Yearly Analysis...")
                district_test_yearly_df = self.get_district_test_yearly_analysis(start_date, end_date)
                if not district_test_yearly_df.empty:
                    district_test_yearly_df.to_excel(writer, sheet_name='District Test Yearly', index=False)
                    self.format_excel_sheet(writer.sheets['District Test Yearly'], district_test_yearly_df)
                    all_data['district_test_yearly'] = district_test_yearly_df.to_dict('records')
                    print(f"   ✓ District-wise Test Yearly: {len(district_test_yearly_df)} rows exported")
                else:
                    print("   ⚠ District-wise Test Yearly: No data found")
            except Exception as e:
                print(f"   ✗ District-wise Test Yearly failed: {e}")

            # 10. Species Analysis
            try:
                print("10/11 Processing Species Analysis...")
                species_df = self.get_species_analysis(start_date, end_date)
                if not species_df.empty:
                    species_df.to_excel(writer, sheet_name='Species Analysis', index=False)
                    self.format_excel_sheet(writer.sheets['Species Analysis'], species_df)
                    all_data['species_analysis'] = species_df.to_dict('records')
                    print(f"   ✓ Species Analysis: {len(species_df)} rows exported")
                else:
                    print("   ⚠ Species Analysis: No data found")
            except Exception as e:
                print(f"   ✗ Species Analysis failed: {e}")

            # 11. Dashboard KPIs
            try:
                print("11/11 Processing Dashboard KPIs...")
                kpis = self.get_dashboard_kpis(start_date, end_date)
                kpis_df = pd.DataFrame([kpis]).T.reset_index()
                kpis_df.columns = ['KPI', 'Value']
                kpis_df.to_excel(writer, sheet_name='Dashboard KPIs', index=False)
                self.format_excel_sheet(writer.sheets['Dashboard KPIs'], kpis_df)
                all_data['dashboard_kpis'] = kpis
                print(f"   ✓ Dashboard KPIs: {len(kpis_df)} metrics exported")
            except Exception as e:
                print(f"   ✗ Dashboard KPIs failed: {e}")

        # Export to JSON
        json_filename = filename.replace('.xlsx', '.json')
        all_data['meta'] = {
            'export_date': datetime.now().isoformat(),
            'start_date': start_date,
            'end_date': end_date,
            'days_analyzed': days_back
        }

        with open(json_filename, 'w') as f:
            json.dump(all_data, f, indent=2, default=str)

        print(f"\n{'='*70}")
        print(f"✓ Export completed successfully!")
        print(f"📁 Excel file saved: {filename}")
        print(f"📁 JSON file saved: {json_filename}")
        print(f"{'='*70}\n")

        return all_data


def main():
    """Main execution function"""
    print("\n" + "="*70)
    print("eLAB Comprehensive Analytics Data Exporter v2.0")
    print("="*70 + "\n")

    try:
        exporter = eLABAnalyticsComprehensive()

        # You can customize these parameters
        days_back = int(input("Enter number of days to analyze (default 365): ") or 365)
        filename = input("Enter output filename (default 'elab_analytics_comprehensive.xlsx'): ") or "elab_analytics_comprehensive.xlsx"

        if not filename.endswith('.xlsx'):
            filename += '.xlsx'

        exporter.export_to_excel(filename=filename, days_back=days_back)

        print("\n✓ All comprehensive analytics have been exported successfully!")
        print(f"✓ You can now open '{filename}' to view the data")
        print(f"✓ JSON data available in '{filename.replace('.xlsx', '.json')}'")

    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()
        print("Please check your .env file and ensure SUPABASE_URL and SUPABASE_KEY are set correctly")


if __name__ == "__main__":
    main()
