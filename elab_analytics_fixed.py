"""
eLAB Analytics Data Exporter - FIXED VERSION
- Fetches ALL data with pagination
- Uses total_test_charge (excludes collection_charges)
- No restrictive date filtering
- Handles 4000+ records properly
"""

import os
import json
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any
import pandas as pd
from supabase import create_client, Client
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

class eLABAnalyticsExporter:
    def __init__(self):
        """Initialize Supabase client"""
        supabase_url = os.getenv("SUPABASE_URL")
        supabase_key = os.getenv("SUPABASE_KEY")

        if not supabase_url or not supabase_key:
            raise ValueError("SUPABASE_URL and SUPABASE_KEY must be set in .env file")

        self.client: Client = create_client(supabase_url, supabase_key)
        print("✓ Connected to Supabase")
        self.cache = {}

    def get_all_data(self, table_name: str, columns: str = "*", batch_size: int = 1000) -> pd.DataFrame:
        """Fetch ALL data from table with pagination"""
        print(f"   Fetching all data from {table_name}...", end=" ")

        all_data = []
        offset = 0

        while True:
            try:
                response = self.client.table(table_name).select(columns).range(offset, offset + batch_size - 1).execute()

                if not response.data:
                    break

                all_data.extend(response.data)
                offset += batch_size

                if len(response.data) < batch_size:
                    break

            except Exception as e:
                print(f"\n   Error at offset {offset}: {e}")
                break

        print(f"Got {len(all_data)} records")
        return pd.DataFrame(all_data)

    def get_date_filter_range(self, days_back: Optional[int] = None):
        """Get date range - if days_back is None, return None (no filter)"""
        if days_back is None:
            return None, None
        end_date = datetime.now()
        start_date = end_date - timedelta(days=days_back)
        return start_date, end_date

    # ========== MODULE 1: District-wise Analysis ==========
    def get_district_analysis(self, days_back: Optional[int] = None) -> pd.DataFrame:
        """District-wise cases, revenue, and trends"""
        print("📊 District-wise Analysis")

        try:
            visits_df = self.get_all_data('visits')
            owners_df = self.get_all_data('owners')

            if visits_df.empty:
                return pd.DataFrame()

            # Convert dates
            visits_df['visit_date'] = pd.to_datetime(visits_df['visit_date'], errors='coerce')

            # Apply date filter if specified
            if days_back:
                start_date, end_date = self.get_date_filter_range(days_back)
                visits_df = visits_df[visits_df['visit_date'] >= start_date]
                print(f"   Filtered to last {days_back} days: {len(visits_df)} visits")

            # Use total_test_charge (excluding collection charges)
            visits_df['revenue'] = pd.to_numeric(visits_df['total_test_charge'], errors='coerce').fillna(0)

            # Merge with owners
            merged = visits_df.merge(
                owners_df[['id', 'district', 'state']],
                left_on='owner_id',
                right_on='id',
                how='left',
                suffixes=('', '_owner')
            )

            # Fill missing districts
            merged['district'] = merged['district'].fillna('Unknown')
            merged['state'] = merged['state'].fillna('Unknown')

            # Aggregate
            summary = merged.groupby(['district', 'state']).agg({
                'id': 'count',
                'owner_id': 'nunique',
                'revenue': ['sum', 'mean']
            }).reset_index()

            summary.columns = ['district', 'state', 'total_cases', 'unique_owners',
                              'total_revenue', 'avg_revenue_per_visit']

            return summary.sort_values('total_cases', ascending=False)

        except Exception as e:
            print(f"   Error: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()

    # ========== MODULE 2-5: Employee Analytics ==========
    def get_employee_performance(self, days_back: Optional[int] = None) -> pd.DataFrame:
        """Employee performance"""
        print("👤 Employee Performance")

        try:
            visits_df = self.get_all_data('visits')

            if visits_df.empty:
                return pd.DataFrame()

            visits_df['visit_date'] = pd.to_datetime(visits_df['visit_date'], errors='coerce')

            if days_back:
                start_date, end_date = self.get_date_filter_range(days_back)
                visits_df = visits_df[visits_df['visit_date'] >= start_date]
                print(f"   Filtered to last {days_back} days: {len(visits_df)} visits")

            # Use total_test_charge
            visits_df['revenue'] = pd.to_numeric(visits_df['total_test_charge'], errors='coerce').fillna(0)

            # Handle employee fields
            visits_df['employee_id'] = visits_df['lab_assistant_id'].fillna(visits_df['field_staff_id'])
            visits_df['employee_name'] = visits_df['lab_staff_name'].fillna(visits_df['field_staff_name'])

            # Remove records without employee
            visits_df = visits_df[visits_df['employee_id'].notna()]

            data = []
            for _, visit in visits_df.iterrows():
                data.append({
                    'employee_id': visit['employee_id'],
                    'employee_name': visit.get('employee_name', 'Unknown'),
                    'visit_id': visit['id'],
                    'visit_date': visit['visit_date'],
                    'revenue': visit['revenue'],
                    'payment_status': bool(visit.get('payment_status', False)),
                    'sample_collected': bool(visit.get('sample_collected', False)),
                    'report_sent': bool(visit.get('report_sent', False))
                })

            return pd.DataFrame(data)

        except Exception as e:
            print(f"   Error: {e}")
            return pd.DataFrame()

    def get_employee_case_comparison(self, days_back: Optional[int] = None) -> pd.DataFrame:
        """Compare case counts across employees"""
        print("📈 Employee Comparison")

        try:
            df = self.get_employee_performance(days_back)

            if df.empty:
                return pd.DataFrame()

            comparison = df.groupby(['employee_id', 'employee_name']).agg({
                'visit_id': 'count',
                'revenue': 'sum',
                'payment_status': 'sum',
                'sample_collected': 'sum',
                'report_sent': 'sum'
            }).reset_index()

            comparison.columns = ['employee_id', 'employee_name', 'total_cases',
                                 'total_revenue', 'payments_received',
                                 'samples_collected', 'reports_sent']

            return comparison.sort_values('total_cases', ascending=False)

        except Exception as e:
            print(f"   Error: {e}")
            return pd.DataFrame()

    def get_employee_test_breakdown(self, days_back: Optional[int] = None) -> pd.DataFrame:
        """Test type breakdown per employee"""
        print("🔬 Employee Test Breakdown")

        try:
            tests_df = self.get_all_data('tests')
            visit_animals_df = self.get_all_data('visit_animals')
            visits_df = self.get_all_data('visits')

            if tests_df.empty or visit_animals_df.empty or visits_df.empty:
                return pd.DataFrame()

            visits_df['visit_date'] = pd.to_datetime(visits_df['visit_date'], errors='coerce')

            if days_back:
                start_date, end_date = self.get_date_filter_range(days_back)
                visits_df = visits_df[visits_df['visit_date'] >= start_date]

            # Merge
            tests_with_visits = tests_df.merge(
                visit_animals_df[['id', 'visit_id']],
                left_on='visit_animal_id',
                right_on='id',
                how='inner',
                suffixes=('', '_va')
            )

            tests_with_emp = tests_with_visits.merge(
                visits_df[['id', 'lab_assistant_id', 'field_staff_id',
                          'lab_staff_name', 'field_staff_name', 'visit_date']],
                left_on='visit_id',
                right_on='id',
                how='inner',
                suffixes=('', '_visit')
            )

            tests_with_emp['employee_id'] = tests_with_emp['lab_assistant_id'].fillna(
                tests_with_emp['field_staff_id']
            )
            tests_with_emp['employee_name'] = tests_with_emp['lab_staff_name'].fillna(
                tests_with_emp['field_staff_name']
            )

            # Use price field from tests table
            tests_with_emp['test_price'] = pd.to_numeric(tests_with_emp['price'], errors='coerce').fillna(0)

            summary = tests_with_emp.groupby([
                'employee_id', 'employee_name', 'test_type', 'test_name'
            ]).agg({
                'id': 'count',
                'test_price': 'sum'
            }).reset_index()

            summary.columns = ['employee_id', 'employee_name', 'test_type',
                              'test_name', 'test_count', 'revenue']

            return summary.sort_values('test_count', ascending=False)

        except Exception as e:
            print(f"   Error: {e}")
            return pd.DataFrame()

    # ========== MODULE 6: Test-based Analysis ==========
    def get_test_analysis(self, days_back: Optional[int] = None) -> pd.DataFrame:
        """Test performance"""
        print("🧪 Test Analysis")

        try:
            tests_df = self.get_all_data('tests')
            visit_animals_df = self.get_all_data('visit_animals')
            visits_df = self.get_all_data('visits')

            if tests_df.empty:
                return pd.DataFrame()

            visits_df['visit_date'] = pd.to_datetime(visits_df['visit_date'], errors='coerce')

            if days_back:
                start_date, end_date = self.get_date_filter_range(days_back)
                visits_df = visits_df[visits_df['visit_date'] >= start_date]

            # Merge
            tests_with_dates = tests_df.merge(
                visit_animals_df[['id', 'visit_id']],
                left_on='visit_animal_id',
                right_on='id',
                how='left',
                suffixes=('', '_va')
            ).merge(
                visits_df[['id', 'visit_date']],
                left_on='visit_id',
                right_on='id',
                how='left',
                suffixes=('', '_visit')
            )

            tests_with_dates['test_price'] = pd.to_numeric(tests_with_dates['price'], errors='coerce').fillna(0)
            tests_with_dates['test_type'] = tests_with_dates['test_type'].fillna('Unknown')
            tests_with_dates['test_name'] = tests_with_dates['test_name'].fillna('Unknown')

            summary = tests_with_dates.groupby(['test_type', 'test_name']).agg({
                'id': 'count',
                'test_price': ['sum', 'mean'],
                'visit_date': ['min', 'max']
            }).reset_index()

            summary.columns = ['test_type', 'test_name', 'test_count',
                              'total_revenue', 'avg_price',
                              'first_test_date', 'last_test_date']

            # Remove timezone info for Excel compatibility
            if 'first_test_date' in summary.columns:
                summary['first_test_date'] = pd.to_datetime(summary['first_test_date']).dt.tz_localize(None)
            if 'last_test_date' in summary.columns:
                summary['last_test_date'] = pd.to_datetime(summary['last_test_date']).dt.tz_localize(None)

            return summary.sort_values('test_count', ascending=False)

        except Exception as e:
            print(f"   Error: {e}")
            return pd.DataFrame()

    # ========== MODULE 7 & 8: Species-wise Analysis ==========
    def get_species_analysis(self, days_back: Optional[int] = None) -> pd.DataFrame:
        """Species-wise cases and revenue"""
        print("🐾 Species Analysis")

        try:
            visit_animals_df = self.get_all_data('visit_animals')
            animals_df = self.get_all_data('animals')
            visits_df = self.get_all_data('visits')

            if visit_animals_df.empty or animals_df.empty or visits_df.empty:
                return pd.DataFrame()

            visits_df['visit_date'] = pd.to_datetime(visits_df['visit_date'], errors='coerce')

            if days_back:
                start_date, end_date = self.get_date_filter_range(days_back)
                visits_df = visits_df[visits_df['visit_date'] >= start_date]

            visits_df['revenue'] = pd.to_numeric(visits_df['total_test_charge'], errors='coerce').fillna(0)

            species_data = visit_animals_df.merge(
                animals_df[['id', 'species', 'breed']],
                left_on='animal_id',
                right_on='id',
                how='left',
                suffixes=('', '_animal')
            ).merge(
                visits_df[['id', 'visit_date', 'revenue']],
                left_on='visit_id',
                right_on='id',
                how='left',
                suffixes=('', '_visit')
            )

            species_data['species'] = species_data['species'].fillna('Unknown')

            summary = species_data.groupby('species').agg({
                'id': 'count',
                'revenue': ['sum', 'mean']
            }).reset_index()

            summary.columns = ['species', 'total_cases', 'total_revenue',
                              'avg_revenue_per_case']

            return summary.sort_values('total_cases', ascending=False)

        except Exception as e:
            print(f"   Error: {e}")
            return pd.DataFrame()

    def get_species_trends(self, days_back: Optional[int] = None) -> pd.DataFrame:
        """Species revenue trends"""
        print("📊 Species Trends")

        try:
            visit_animals_df = self.get_all_data('visit_animals')
            animals_df = self.get_all_data('animals')
            visits_df = self.get_all_data('visits')

            if visit_animals_df.empty or animals_df.empty or visits_df.empty:
                return pd.DataFrame()

            visits_df['visit_date'] = pd.to_datetime(visits_df['visit_date'], errors='coerce')

            if days_back:
                start_date, end_date = self.get_date_filter_range(days_back)
                visits_df = visits_df[visits_df['visit_date'] >= start_date]

            visits_df['revenue'] = pd.to_numeric(visits_df['total_test_charge'], errors='coerce').fillna(0)

            species_trends = visit_animals_df.merge(
                animals_df[['id', 'species']],
                left_on='animal_id',
                right_on='id',
                how='left',
                suffixes=('', '_animal')
            ).merge(
                visits_df[['id', 'visit_date', 'revenue']],
                left_on='visit_id',
                right_on='id',
                how='left',
                suffixes=('', '_visit')
            )

            species_trends['species'] = species_trends['species'].fillna('Unknown')
            species_trends = species_trends[species_trends['visit_date'].notna()]
            species_trends['year_month'] = species_trends['visit_date'].dt.to_period('M').astype(str)

            trends = species_trends.groupby(['species', 'year_month']).agg({
                'revenue': ['sum', 'count']
            }).reset_index()

            trends.columns = ['species', 'year_month', 'revenue', 'case_count']

            return trends

        except Exception as e:
            print(f"   Error: {e}")
            return pd.DataFrame()

    # ========== MODULE 9: Month-wise Analysis ==========
    def get_monthly_analysis(self, days_back: Optional[int] = None) -> pd.DataFrame:
        """Month-wise revenue and cases analysis"""
        print("📅 Monthly Analysis")

        try:
            visits_df = self.get_all_data('visits')

            if visits_df.empty:
                return pd.DataFrame()

            visits_df['visit_date'] = pd.to_datetime(visits_df['visit_date'], errors='coerce')
            visits_df = visits_df[visits_df['visit_date'].notna()]

            if days_back:
                start_date, end_date = self.get_date_filter_range(days_back)
                visits_df = visits_df[visits_df['visit_date'] >= start_date]

            visits_df['revenue'] = pd.to_numeric(visits_df['total_test_charge'], errors='coerce').fillna(0)
            visits_df['year_month'] = visits_df['visit_date'].dt.to_period('M').astype(str)

            monthly = visits_df.groupby('year_month').agg({
                'id': 'count',
                'revenue': ['sum', 'mean'],
                'owner_id': 'nunique',
                'payment_status': 'sum',
                'sample_collected': 'sum',
                'report_sent': 'sum'
            }).reset_index()

            monthly.columns = ['month', 'total_cases', 'total_revenue', 'avg_revenue',
                              'unique_customers', 'payments', 'samples_collected', 'reports_sent']

            monthly = monthly.sort_values('month')
            return monthly

        except Exception as e:
            print(f"   Error: {e}")
            return pd.DataFrame()

    # ========== MODULE 10: Year-wise Analysis ==========
    def get_yearly_analysis(self, days_back: Optional[int] = None) -> pd.DataFrame:
        """Year-wise revenue and cases analysis"""
        print("📅 Yearly Analysis")

        try:
            visits_df = self.get_all_data('visits')

            if visits_df.empty:
                return pd.DataFrame()

            visits_df['visit_date'] = pd.to_datetime(visits_df['visit_date'], errors='coerce')
            visits_df = visits_df[visits_df['visit_date'].notna()]

            if days_back:
                start_date, end_date = self.get_date_filter_range(days_back)
                visits_df = visits_df[visits_df['visit_date'] >= start_date]

            visits_df['revenue'] = pd.to_numeric(visits_df['total_test_charge'], errors='coerce').fillna(0)
            visits_df['year'] = visits_df['visit_date'].dt.year

            yearly = visits_df.groupby('year').agg({
                'id': 'count',
                'revenue': ['sum', 'mean'],
                'owner_id': 'nunique',
                'payment_status': 'sum',
                'sample_collected': 'sum',
                'report_sent': 'sum'
            }).reset_index()

            yearly.columns = ['year', 'total_cases', 'total_revenue', 'avg_revenue',
                             'unique_customers', 'payments', 'samples_collected', 'reports_sent']

            yearly = yearly.sort_values('year')
            return yearly

        except Exception as e:
            print(f"   Error: {e}")
            return pd.DataFrame()

    # ========== MODULE 11: Dashboard KPIs ==========
    def get_dashboard_kpis(self, days_back: Optional[int] = None) -> Dict[str, Any]:
        """Dashboard KPIs"""
        print("📊 Dashboard KPIs")

        try:
            visits_df = self.get_all_data('visits')

            if visits_df.empty:
                return {}

            visits_df['visit_date'] = pd.to_datetime(visits_df['visit_date'], errors='coerce')

            if days_back:
                start_date, end_date = self.get_date_filter_range(days_back)
                visits_df = visits_df[visits_df['visit_date'] >= start_date]

            visits_df['revenue'] = pd.to_numeric(visits_df['total_test_charge'], errors='coerce').fillna(0)

            kpis = {
                'total_cases': int(len(visits_df)),
                'total_revenue': float(visits_df['revenue'].sum()),
                'avg_revenue_per_case': float(visits_df['revenue'].mean()) if len(visits_df) > 0 else 0,
                'payment_completion_rate': float(visits_df['payment_status'].sum() / len(visits_df) * 100) if len(visits_df) > 0 else 0,
                'sample_collection_rate': float(visits_df['sample_collected'].sum() / len(visits_df) * 100) if len(visits_df) > 0 else 0,
                'report_sent_rate': float(visits_df['report_sent'].sum() / len(visits_df) * 100) if len(visits_df) > 0 else 0,
                'unique_customers': int(visits_df['owner_id'].nunique()) if 'owner_id' in visits_df else 0,
            }

            users_df = self.get_all_data('users')
            if not users_df.empty:
                kpis['total_employees'] = int(len(users_df))

            return kpis

        except Exception as e:
            print(f"   Error: {e}")
            return {}

    # ========== Excel Export ==========
    def format_excel_sheet(self, worksheet, df: pd.DataFrame):
        """Apply formatting"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for idx, column in enumerate(df.columns, 1):
            column_letter = get_column_letter(idx)
            max_length = max(df[column].astype(str).apply(len).max(), len(str(column)))
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    def export_to_excel(self, filename: str = "elab_analytics_full.xlsx", days_back: Optional[int] = None):
        """Export all analytics"""
        print(f"\n{'='*60}")
        print(f"eLAB Analytics Export - FULL DATA")
        print(f"{'='*60}\n")

        if days_back:
            start_date, end_date = self.get_date_filter_range(days_back)
            print(f"📅 Filtering to last {days_back} days\n")
        else:
            print(f"📅 Exporting ALL DATA (no date filter)\n")

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:

            modules = [
                ('District Analysis', lambda: self.get_district_analysis(days_back)),
                ('Employee Comparison', lambda: self.get_employee_case_comparison(days_back)),
                ('Employee Test Breakdown', lambda: self.get_employee_test_breakdown(days_back)),
                ('Test Analysis', lambda: self.get_test_analysis(days_back)),
                ('Species Analysis', lambda: self.get_species_analysis(days_back)),
                ('Species Trends', lambda: self.get_species_trends(days_back)),
                ('Monthly Analysis', lambda: self.get_monthly_analysis(days_back)),
                ('Yearly Analysis', lambda: self.get_yearly_analysis(days_back)),
            ]

            for sheet_name, func in modules:
                try:
                    df = func()
                    if not df.empty:
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        self.format_excel_sheet(writer.sheets[sheet_name], df)
                        print(f"✓ {sheet_name}: {len(df)} rows")
                    else:
                        print(f"⚠ {sheet_name}: No data")
                except Exception as e:
                    print(f"✗ {sheet_name}: {e}")

            # KPIs
            try:
                kpis = self.get_dashboard_kpis(days_back)
                if kpis:
                    kpis_df = pd.DataFrame([kpis]).T.reset_index()
                    kpis_df.columns = ['KPI', 'Value']
                    kpis_df.to_excel(writer, sheet_name='Dashboard KPIs', index=False)
                    self.format_excel_sheet(writer.sheets['Dashboard KPIs'], kpis_df)
                    print(f"✓ Dashboard KPIs: {len(kpis_df)} metrics")
            except Exception as e:
                print(f"✗ Dashboard KPIs: {e}")

        print(f"\n{'='*60}")
        print(f"✓ Export completed!")
        print(f"📁 File: {filename}")
        print(f"{'='*60}\n")

    def export_to_json(self, filename: str = "elab_analytics_data.json", days_back: Optional[int] = None):
        """Export to JSON for dashboard"""
        print(f"Exporting to JSON: {filename}")

        if days_back:
            start_date, end_date = self.get_date_filter_range(days_back)
            date_info = {
                'start_date': start_date.strftime('%Y-%m-%d'),
                'end_date': end_date.strftime('%Y-%m-%d')
            }
        else:
            date_info = {
                'start_date': 'ALL',
                'end_date': 'ALL'
            }

        data = {
            'meta': {
                **date_info,
                'generated_at': datetime.now().isoformat()
            }
        }

        try:
            data['district_analysis'] = self.get_district_analysis(days_back).to_dict('records')
        except:
            data['district_analysis'] = []

        try:
            data['employee_comparison'] = self.get_employee_case_comparison(days_back).to_dict('records')
        except:
            data['employee_comparison'] = []

        try:
            data['employee_tests'] = self.get_employee_test_breakdown(days_back).to_dict('records')
        except:
            data['employee_tests'] = []

        try:
            data['test_analysis'] = self.get_test_analysis(days_back).to_dict('records')
        except:
            data['test_analysis'] = []

        try:
            data['species_analysis'] = self.get_species_analysis(days_back).to_dict('records')
        except:
            data['species_analysis'] = []

        try:
            data['species_trends'] = self.get_species_trends(days_back).to_dict('records')
        except:
            data['species_trends'] = []

        try:
            data['monthly_analysis'] = self.get_monthly_analysis(days_back).to_dict('records')
        except:
            data['monthly_analysis'] = []

        try:
            data['yearly_analysis'] = self.get_yearly_analysis(days_back).to_dict('records')
        except:
            data['yearly_analysis'] = []

        try:
            data['dashboard_kpis'] = self.get_dashboard_kpis(days_back)
        except:
            data['dashboard_kpis'] = {}

        with open(filename, 'w') as f:
            json.dump(data, f, indent=2, default=str)

        print(f"✓ JSON exported\n")
        return data


def main():
    print("\n" + "="*60)
    print("eLAB Analytics Exporter - FIXED VERSION")
    print("="*60 + "\n")

    try:
        exporter = eLABAnalyticsExporter()

        print("\nOptions:")
        print("1. Export ALL data (no date filter)")
        print("2. Export data for specific time period")
        choice = input("\nYour choice (1 or 2): ").strip()

        if choice == "1":
            days_back = None
            print("\n→ Exporting ALL DATA\n")
        else:
            days_back = int(input("Enter number of days (e.g., 30, 90, 365): ") or 365)

        filename = input("Excel filename (default 'elab_analytics_full.xlsx'): ") or "elab_analytics_full.xlsx"
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'

        exporter.export_to_excel(filename=filename, days_back=days_back)
        exporter.export_to_json(filename="elab_analytics_data.json", days_back=days_back)

        print("\n✓ Export complete!")
        print(f"✓ Excel: {filename}")
        print(f"✓ JSON: elab_analytics_data.json")
        print("\n→ Open elab_analytics_dashboard.html to view the dashboard\n")

    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
